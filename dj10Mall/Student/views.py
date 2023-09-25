from django.shortcuts import render, HttpResponse, redirect
from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Student, StudentDetail, StuCourse, Clas
import os
from openpyxl import load_workbook
from django.contrib import auth

from .serializers import ClssSerializer


class ClssView(APIView):
    def get(self, request):
        # 通过ORM操作获取所有分类数据
        queryset = Clas.objects.all()
        # 利用序列化器去序列化我们的数据
        ser_obj = ClssSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
        # 返回
        return Response(ser_obj.data)

    def post(self, request):
        print('post', request.data)
        name = request.data.get('name', '')
        res = Clas.objects.create(name=name)
        if res:
            return Response('ok')
        return Response('')

    def put(self, request):
        print('put:', request.data)
        id = request.data.get('id', '')
        name = request.data.get('name', '')
        res = Clas.objects.filter(id=int(id)).update(name=name)
        if res:
            return Response('ok')
        return Response('')

    def delete(self, request):
        print('delete:', request.GET)
        id = request.GET.get('id', '')
        res = Clas.objects.filter(id=int(id)).delete()
        if res:
            return Response('ok')
        return Response('')


def index(request):
    # （1）视图函数单独验证
    # if request.user.username:
    #     # 获取所有的学生数据
    #     student_list = Student.objects.all()
    #
    #     return render(request, "student/index.html", {"student_list": student_list})
    # else:
    #     return redirect("/login")

    # （2）中间件判断
    # 获取所有的学生数据
    student_list = Student.objects.all()

    return render(request, "student/index.html", {"student_list": student_list})


def add_student(request):
    if request.method == "GET":

        class_list = Clas.objects.all()
        course_list = StuCourse.objects.all()

        return render(request, "student/add_stu.html", {"class_list": class_list, "course_list": course_list})

    else:
        # 获取客户端数据
        # <QueryDict: {'name': ['alvin'], 'age': ['22'], 'sex': ['2'], 'birthday': ['2021-10-26'], 'clas_id': ['8']}>
        print(request.POST)
        # 添加数据到数据库
        # 方式1
        # name = request.POST.get("user")
        # age = request.POST.get("age")
        # sex = request.POST.get("sex")
        # birthday = request.POST.get("birthday")
        # clas_id = request.POST.get("clas_id")
        #
        # Student.objects.create(name=name,age=age,sex=sex,birthday=birthday,clas_id=clas_id)
        # 方式2
        # 前提： 客户端form表单的name值要与models中的字段保持一致
        stu = Student.objects.create(**request.POST.dict())

        return redirect("/student/")


def delete_student(request, del_id):
    student = Student.objects.get(pk=del_id)

    if request.method == "GET":
        return render(request, "student/delete_stu.html", {"student": student})

    else:
        student.delete()
        return redirect("/student/")


def edit_student(request, edit_id):
    edit_stu = Student.objects.get(pk=edit_id)
    if request.method == "GET":

        class_list = Clas.objects.all()
        course_list = StuCourse.objects.all()
        return render(request, "student/edit_stu.html",
                      {"edit_stu": edit_stu, "class_list": class_list, "course_list": course_list})
    else:
        print("request.POST", request.POST)
        course_id_list = request.POST.getlist("course_id_list")
        # 获取客户端数据
        data = request.POST.dict()
        print("data", data)
        # 删除并获取课程id列表
        data.pop("course_id_list")
        # 更新除了多对多以外的数据
        Student.objects.filter(pk=edit_id).update(**data)
        # 多对多关系的重置
        edit_stu.courses.set(course_id_list)
        return redirect("/student/")


def elective(request):
    if request.method == "GET":
        course_list = StuCourse.objects.all()
        return render(request, "student/course.html", {"course_list": course_list})

    else:
        print(request.POST)
        course_id_list = request.POST.getlist("course_id_list")
        stu_id = request.user.stu_id
        stu = Student.objects.get(pk=stu_id)
        stu.courses.set(course_id_list)

        return redirect("/student/")


def search(request):
    cate = request.GET.get("cate")
    key_word = request.GET.get("key_word")

    if cate == "name":
        student_list = Student.objects.filter(name__contains=key_word)

    elif cate == "class":
        student_list = Student.objects.filter(clas__name=key_word)
    else:
        student_list = []

    return render(request, "student/index.html", {"student_list": student_list, "key_word": key_word})


def stu_excel(request):
    excel_stus = request.FILES.get("excel_stus")
    print(excel_stus)
    print(excel_stus.name)

    # (1) 将上传文件下载到服务器某个文件夹下

    path = os.path.join("media", "files", excel_stus.name)
    with open(path, "wb") as f:

        for line in excel_stus:
            f.write(line)

    # (2) 通过python操作excel表格
    wb = load_workbook(path)
    print(wb.sheetnames)

    work_sheet = wb.worksheets[0]

    stu_list = []
    for line in work_sheet.iter_rows(min_row=3):
        # print("line",line)
        # for cell in line:
        #     print(cell.value)

        if not line[0].value:
            break

        sd = StudentDetail.objects.create(tel=line[4].value, addr=line[5].value)
        class_id = Clas.objects.get(name=line[6].value).id

        if line[2].value == "男":
            sex = 1
        elif line[2].value == "女":
            sex = 0
        else:
            sex = 2

        stu = Student(name=line[0].value,
                      age=line[1].value,
                      sex=sex,
                      birthday=line[3].value,

                      clas_id=class_id,
                      stu_detail=sd
                      )

        stu_list.append(stu)

    Student.objects.bulk_create(stu_list)

    return redirect("/student/")


def login(request):
    if request.method == "GET":

        return render(request, "student/login.html")

    else:

        user = request.POST.get("loginUsername")
        pwd = request.POST.get("loginPassword")

        # 使用了用户认证组件
        user = auth.authenticate(username=user, password=pwd)
        if user:
            # 验证成功

            # 写session: request.session["user_id"] =user.id
            auth.login(request, user)
            return redirect("/")

        else:
            # 验证失败
            return redirect("/login")


def logout(request):
    auth.logout(request)

    return redirect("/login/")
