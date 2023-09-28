from rest_framework.views import APIView
from rest_framework.response import Response
from Student.models import StudentDetail, ClassRoom, Teacher, Student
import os
import json
from openpyxl import load_workbook
from django.db import connection

from Student.serializers.schoolBaseSerializer import TeacherSerializer, \
    StudentDetailSerializer, ClassRoomSerializer, StudentSerializer


class SchoolTeacherView(APIView):
    def get(self, request):
        print('request.GET:', request.GET)
        name = request.GET.get('name', '')
        if name:
            queryset = Teacher.objects.filter(name__contains=name)
            ser_obj = TeacherSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
            # 返回
            return Response(ser_obj.data)

        # 通过ORM操作获取所有分类数据
        queryset = Teacher.objects.all()
        # 利用序列化器去序列化我们的数据
        ser_obj = TeacherSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
        # 返回
        return Response(ser_obj.data)

    def post(self, request):
        print('post-data', request.data)
        name = request.data.get('name', '')
        res = Teacher.objects.create(name=name)
        if res:
            return Response('ok')
        return Response('')

    def put(self, request):
        print('put:', request.data)
        id = request.data.get('id', '')
        name = request.data.get('name', '')
        res = Teacher.objects.filter(id=int(id)).update(name=name)
        if res:
            return Response('ok')
        return Response('')

    def delete(self, request):
        print('delete111:', request.GET)
        id = request.GET.get('del_list', '')
        if id:
            res = Teacher.objects.filter(id=int(id)).delete()  # 用id__in 来拿取数据 紧接着删除
            if res:
                return Response('ok')

        # TODO 批量删除 待完善
        # del_list = request.GET.get('del_list')  # 获取前台传过来的列表
        # print(type del_list)
        # if del_list:
        #     print('del_list:', del_list)
        #     res = Clas.objects.filter(id__in=del_list).delete()  # 用id__in 来拿取数据 紧接着删除
        #
        #     if res:
        #         return Response('ok')

        return Response('')


class SchoolStudentDetailView(APIView):
    def get(self, request):
        print('request.GET:', request.GET)
        name = request.GET.get('name', '')
        if name:
            queryset = StudentDetail.objects.filter(name__contains=name)
            ser_obj = StudentDetailSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
            # 返回
            return Response(ser_obj.data)

        # 通过ORM操作获取所有分类数据
        queryset = StudentDetail.objects.all()
        # 利用序列化器去序列化我们的数据
        ser_obj = StudentDetailSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
        # 返回
        return Response(ser_obj.data)

    def post(self, request):
        print('post-data', request.data)
        name = request.data.get('name', '')
        res = StudentDetail.objects.create(name=name)
        if res:
            return Response('ok')
        return Response('')

    def put(self, request):
        print('put:', request.data)
        id = request.data.get('id', '')
        name = request.data.get('name', '')
        res = StudentDetail.objects.filter(id=int(id)).update(name=name)
        if res:
            return Response('ok')
        return Response('')

    def delete(self, request):
        print('delete111:', request.GET)
        id = request.GET.get('del_list', '')
        if id:
            res = StudentDetail.objects.filter(id=int(id)).delete()  # 用id__in 来拿取数据 紧接着删除
            if res:
                return Response('ok')

        # TODO 批量删除 待完善
        # del_list = request.GET.get('del_list')  # 获取前台传过来的列表
        # print(type del_list)
        # if del_list:
        #     print('del_list:', del_list)
        #     res = Clas.objects.filter(id__in=del_list).delete()  # 用id__in 来拿取数据 紧接着删除
        #
        #     if res:
        #         return Response('ok')

        return Response('')


class SchoolStudentView(APIView):
    def get(self, request):
        print('request.GET:', request.GET)
        name = request.GET.get('name', '')
        if name:
            queryset = Student.objects.filter(name__contains=name)
            ser_obj = StudentSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
            # detail_id = ser_obj.stu_detail_id
            print('detail_id-name', ser_obj)
            # 返回
            return Response(ser_obj.data)

        # 通过ORM操作获取所有分类数据
        # queryset = Student.objects.all()
        # # 利用序列化器去序列化我们的数据
        # ser_obj = StudentSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
        # # detail_id = ser_obj.stu_detail_id
        # print('detail_id-all', ser_obj)

        # 参考原生查询 https://cloud.tencent.com/developer/article/1750639
        # 原声sql 查询
        cursor = connection.cursor()
        cursor.execute('SELECT * FROM db_student LEFT JOIN db_stu_detail ON db_student.stu_detail_id = db_stu_detail.id')
        raw_all = cursor.fetchall()
        # print(99,  raw_all)
        student_list = []
        for key in raw_all:
            print(key[1])
            student_list.append({
                "id": key[0],
                "name": key[1],
                "mobile": key[2],
                "nickname": key[5],
                "avatar": '/media' + str(key[6]),
                "family_addr": key[7],
                "birthday": str(key[8]),
                "father_name": key[9],
                "mother_name": key[10],
                "sex": key[11],
            })
        print(88, student_list)
        cursor.close()

        cur_list = [{'name': 66}, {'name': 77}]
        print(77, cur_list)
        # 返回
        # return Response(json.dumps(student_list))
        return Response(student_list)

    def post(self, request):
        print('post-data', request.data)
        name = request.data.get('name', '')
        res = Student.objects.create(name=name)
        if res:
            return Response('ok')
        return Response('')

    def put(self, request):
        print('put:', request.data)
        id = request.data.get('id', '')
        name = request.data.get('name', '')
        res = Student.objects.filter(id=int(id)).update(name=name)
        if res:
            return Response('ok')
        return Response('')

    def delete(self, request):
        print('delete111:', request.GET)
        id = request.GET.get('del_list', '')
        if id:
            res = Student.objects.filter(id=int(id)).delete()  # 用id__in 来拿取数据 紧接着删除
            if res:
                return Response('ok')

        # TODO 批量删除 待完善
        # del_list = request.GET.get('del_list')  # 获取前台传过来的列表
        # print(type del_list)
        # if del_list:
        #     print('del_list:', del_list)
        #     res = Clas.objects.filter(id__in=del_list).delete()  # 用id__in 来拿取数据 紧接着删除
        #
        #     if res:
        #         return Response('ok')

        return Response('')


class SchoolClassRoomView(APIView):
    def get(self, request):
        print('request.GET:', request.GET)
        name = request.GET.get('name', '')
        if name:
            queryset = ClassRoom.objects.filter(name__contains=name)
            ser_obj = ClassRoomSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
            # 返回
            return Response(ser_obj.data)

        # 通过ORM操作获取所有分类数据
        queryset = ClassRoom.objects.all()
        # 利用序列化器去序列化我们的数据
        ser_obj = ClassRoomSerializer(queryset, many=True)  # 返回多条数据时，加上many=True
        # 返回
        return Response(ser_obj.data)

    def post(self, request):
        print('post-data', request.data)
        name = request.data.get('name', '')
        res = ClassRoom.objects.create(name=name)
        if res:
            return Response('ok')
        return Response('')

    def put(self, request):
        print('put:', request.data)
        id = request.data.get('id', '')
        name = request.data.get('name', '')
        res = ClassRoom.objects.filter(id=int(id)).update(name=name)
        if res:
            return Response('ok')
        return Response('')

    def delete(self, request):
        print('delete111:', request.GET)
        id = request.GET.get('del_list', '')
        if id:
            res = ClassRoom.objects.filter(id=int(id)).delete()  # 用id__in 来拿取数据 紧接着删除
            if res:
                return Response('ok')

        # TODO 批量删除 待完善
        # del_list = request.GET.get('del_list')  # 获取前台传过来的列表
        # print(type del_list)
        # if del_list:
        #     print('del_list:', del_list)
        #     res = Clas.objects.filter(id__in=del_list).delete()  # 用id__in 来拿取数据 紧接着删除
        #
        #     if res:
        #         return Response('ok')

        return Response('')


class SchoolBachExcel(APIView):
    def post(self, request):
        excel_stus = request.FILES.get("excel_file")
        # print(excel_stus)
        # print(excel_stus.name)

        # (1) 将上传文件下载到服务器某个文件夹下[确保文件夹已经存在]
        path = os.path.join("media", "files", excel_stus.name)
        with open(path, "wb") as f:
            for line in excel_stus:
                f.write(line)

        # (2) 通过python操作excel表格
        wb = load_workbook(path)
        print(wb.sheetnames)  # 获取所有表的名字

        # 0 db_class_room
        # 1 db_teacher
        # 2 db_stu_detail
        # work_sheet = wb.worksheets[0]  # excel 文件中，序列号第四张表（从左到右）
        # class_room_list = []
        # for line in work_sheet.iter_rows(min_row=3):
        #     # print("line--读取文件信息", line)
        #     # for cell in line:
        #     #     print(cell.value)
        #     if not line[0].value:
        #         break
        #     # print('capacity_num:', line[2].value)
        #     item_class_room = ClassRoom(
        #         name=line[0].value,
        #         room_addr=line[1].value,
        #         capacity_num=line[2].value,
        #         isMedia_room=line[3].value,
        #     )

        #     class_room_list.append(item_class_room)
        # ClassRoom.objects.bulk_create(class_room_list)

        # 0 db_class_room
        # 1 db_teacher
        # 2 db_stu_detail
        work_sheet_tea = wb.worksheets[1]  # excel 文件中，序列号第四张表（从左到右）
        tea_list = []
        for line in work_sheet_tea.iter_rows(min_row=3):  # 第一行 row = 1
            if not line[0].value:
                break

            item_tea = Teacher(
                name=line[0].value,
                sex=line[1].value,
                mobile=line[2].value,
                # avatar=line[3].value,
                brief=line[4].value,
            )

        #     tea_list.append(item_tea)
        # Teacher.objects.bulk_create(tea_list)

        # 0 db_class_room
        # 1 db_teacher
        # 2 db_stu_detail
        # work_sheet_stu_detail = wb.worksheets[2]  # excel 文件中，序列号第四张表（从左到右）
        # stu_detail_list = []
        # for line in work_sheet_stu_detail.iter_rows(min_row=3):
        #     for cell in line:
        #         print(cell.value)
        #     if not line[0].value:
        #         break
        #     print('birthday', line[3].value)
        #     item_stu_detail = StudentDetail(
        #         nickname=line[0].value,
        #         # avatar=line[1].value,
        #         family_addr=line[2].value,
        #         birthday=line[3].value,
        #         father_name=line[4].value,
        #         mother_name=line[5].value,
        #         sex=int(line[6].value),
        #     )
        #
        #     stu_detail_list.append(item_stu_detail)
        # StudentDetail.objects.bulk_create(stu_detail_list)

        work_sheet_stu = wb.worksheets[3]  # excel 文件中，序列号第四张表（从左到右）
        stu_list = []
        for line in work_sheet_stu.iter_rows(min_row=3):
            # for cell in line:
            #     print(cell.value)
            if not line[0].value:
                break
            print('name', line[0].value)
            item_stu = Student(
                name=line[0].value,
                mobile=line[1].value,
                stu_detail_id=line[2].value
            )

            stu_list.append(item_stu)
        Student.objects.bulk_create(stu_list)

        return Response('ok')
